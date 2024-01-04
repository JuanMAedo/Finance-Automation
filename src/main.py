from ctypes import alignment
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from parser_datetime import *
from dotenv import load_dotenv

#SECRETS LOADINGS AT VARIABLES
load_dotenv()
input_file_path = os.getenv('INCOME_EXPENSE_RECORD')
monthly_income_expense_report = os.getenv('MONTHLY_REPORT_NAME')
table_start_colum = os.getenv('START_COLUMN')
table_finish_column = os.getenv('FINISH_COLUMN')

def create_monthly_column_sheets(month_sheet,title, init_column,text_column,color_title,color_column):
    # PARSER INIT_COLUMN TO ITERATIONS
    column_letter = init_column[0]  
    column_number = column_index_from_string(column_letter)  
    row_number = int(init_column[1:])  
    
    month_sheet[init_column] = f"{title}"
    cell_B2 = month_sheet[init_column]
    cell_B2.alignment = Alignment(horizontal='center')
    cell_B2.fill = PatternFill(start_color=color_title, end_color=color_title, fill_type="solid")  # Color morado claro
    cell_B2.font = Font(bold=True, name='Calibri')
    
    for i in range(5):
        current_row = row_number + i
        cell = month_sheet.cell(row=current_row + 1, column=column_number)
        week_text = f"{text_column} {i + 1}"
        
        cell.value = week_text
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=color_column, end_color=color_column, fill_type="solid")  # Color naranja apagado
        cell.font = Font(bold=True, name='Calibri') 



# VERIFY INPUT FILE
if os.path.exists(input_file_path) and os.path.splitext(input_file_path)[1] in ('.xls', '.xlsx'):
        print("File it's correct")         
else:
    print("The pathing '"'{}'"' or the file type are incorrect.".format(input_file_path))

# OPEN AND READ THE INPUT EXCEL
expense_income_read = load_workbook(input_file_path)
excel_date = expense_income_read.active[table_start_colum].value
# Verify the date
if isinstance(excel_date, datetime.datetime):
    #Parses the date    
    year = excel_date.year
    month = excel_date.month
    day = excel_date.day

#print(excel_date.weekday()) # 0 es Lunes, 6 es domingo --> Para el parseado semanal de las hojas
#print(excel_date.month)



#CREATE OR COMPLETE THE ANNUAL MONTHLY I&E REPORT
file_name = f"{monthly_income_expense_report} {year}.xlsx"
print(file_name)
if os.path.exists(file_name):
    inc_exp_excel = load_workbook(file_name)
    hoja = inc_exp_excel.active

    
    
    # ... 
    
else:
    inc_exp_excel = Workbook()
    inc_exp_excel.create_sheet("DASHBOARD")
    inc_exp_excel.create_sheet("CATEGORY")
    # CREATE MONTHLY SHEETS
    for month_number, month_name in month_to_name.items():
        month_sheet = inc_exp_excel.create_sheet(month_name)
        inc_exp_excel.active = month_sheet
        create_monthly_sheets(month_sheet,"INCOME", 'B10', "Week",'24D124','D78740')          
        create_monthly_sheets(month_sheet,"EXPENSE", 'B2', "Week",'C06FCA','D78740')
                   
    inc_exp_excel.remove(inc_exp_excel["Sheet"])
    # ... 
    
inc_exp_excel.save(file_name)


